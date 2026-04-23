from __future__ import annotations

import functools
import os
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import (
    Flask, abort, jsonify, redirect, render_template,
    request, session, url_for,
)

import json
import tempfile

from config import APP_ENV
from csv_analyzer import analyze, iter_data_rows, normalize_status
from db import get_connection, init_db
from merge_ops import merge_client_into
from name_matcher import find_client as fuzzy_find_client, is_business, parse_name, _all_clients_cache
from normalizer import normalize_date, normalize_currency, normalize_string
from utils import now

app = Flask(__name__)

# Secret key for signing session cookies.
# Set TAXOPS_SECRET env-var in production; a random fallback is fine for dev.
app.secret_key = os.environ.get("TAXOPS_SECRET", os.urandom(24))

# Login credentials — override via environment variables.
_LOGIN_USER = os.environ.get("TAXOPS_USER", "info")
_LOGIN_PASS = os.environ.get("TAXOPS_PASS", "2703Tax")


def privacy_mode_enabled() -> bool:
    return bool(session.get("privacy_mode"))


def _mask_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        return "XXXXX" if value.strip() else value
    return "XXXXX"


def _mask_return_payload(payload: dict) -> dict:
    masked = dict(payload)
    for key in (
        "last_name", "first_name", "display_name", "name_full",
        "referred_by", "ssn_last4", "zelle_or_check_ref",
        "cash_or_qpay_ref", "receipt_number",
    ):
        if key in masked:
            masked[key] = _mask_value(masked.get(key))
    return masked


def _mask_client_payload(payload: dict) -> dict:
    masked = dict(payload)
    for key in (
        "last_name", "first_name", "display_name", "ssn_last4",
        "spouse_last_name", "spouse_first_name",
        "taxpayer_phone", "taxpayer_cell", "taxpayer_work_phone",
        "spouse_cell", "spouse_work_phone",
        "taxpayer_email", "spouse_email",
        "address", "referred_by",
    ):
        if key in masked:
            masked[key] = _mask_value(masked.get(key))
    return masked


def login_required(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


@app.after_request
def _security_headers(response):
    """Add basic security headers — this app is internal-only."""
    response.headers["X-Frame-Options"]        = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"]        = "same-origin"
    response.headers["Cache-Control"]          = "no-store"
    return response

# ── Workflow constants ────────────────────────────────────────────────────────

STATUS_FLOW = ["PROCESSING", "FINALIZE", "PICKUP", "EFILE READY", "LOG OUT", "REJECTED"]

STATUS_BADGE = {
    "PROCESSING":  "bg-sky-50 text-sky-700 border-sky-200",
    "FINALIZE":    "bg-yellow-50 text-yellow-700 border-yellow-200",
    "PICKUP":      "bg-teal-50 text-teal-700 border-teal-200",
    "EFILE READY": "bg-indigo-50 text-indigo-700 border-indigo-200",
    "LOG OUT":     "bg-slate-100 text-slate-500 border-slate-200",
    "REJECTED":    "bg-red-50 text-red-700 border-red-200",
}

STATUS_DOT = {
    "PROCESSING":  "dot-amber",   # sky blue (#0ea5e9)
    "FINALIZE":    "dot-orange",  # yellow (#eab308)
    "PICKUP":      "dot-teal",
    "EFILE READY": "dot-indigo",
    "LOG OUT":     "dot-slate",
    "REJECTED":    "dot-red",
}

# When advancing to these statuses, auto-stamp the corresponding date field
# only if it hasn't been set yet.
STATUS_DATE_STAMP = {
    "PICKUP":  "pickup_date",   # client called in to sign
    "LOG OUT": "logout_date",   # accepted, case closed
}

# Operational risk thresholds
LATE_INTAKE_MONTH = 4
LATE_INTAKE_DAY = 1
SLOW_CYCLE_DAYS = 21

# Fields that live in the returns table and may be edited via /api/return/<id>/field
RETURN_EDITABLE = {
    "processor", "verified", "email_marker",
    "intake_date", "date_emailed", "pickup_date", "logout_date", "updated_date",
    "efile_date", "ack_date",
    "is_amended", "has_w7", "is_extension",
    "transfer_flag", "transfer_2025_flag", "transfer_2026_flag",
}

# Fields that live in the clients table
CLIENT_EDITABLE = {"display_name", "referred_by", "referral_flag"}

# Fields that live in the payments table
PAYMENT_EDITABLE = {
    "total_fee", "fee_paid", "receipt_number",
    "cc_fee", "zelle_or_check_ref", "cash_or_qpay_ref",
    "bank_deposit", "refund_amount",
}

# ── SQL fragment shared by all queries ───────────────────────────────────────

_SELECT = """
SELECT
    r.id, r.log_number, r.tax_year, r.client_status, r.processor,
    r.verified, r.intake_date, r.date_emailed, r.pickup_date,
    r.logout_date, r.updated_date, r.email_marker,
    r.is_amended, r.has_w7, r.is_extension,
    r.transfer_flag, r.transfer_2025_flag, r.transfer_2026_flag,
    r.efile_date, r.ack_date, r.drake_status_raw,
    r.created_at, r.updated_at,
    c.id   AS client_id,
    c.last_name, c.first_name, c.display_name,
    c.referral_flag, c.referred_by,
    p.id   AS payment_id,
    p.total_fee, p.fee_paid, p.receipt_number,
    p.cc_fee, p.zelle_or_check_ref, p.cash_or_qpay_ref,
    p.refund_amount, p.bank_deposit,
    rf.form_1040, rf.sched_a_d, rf.sched_c, rf.sched_e,
    rf.form_1120, rf.form_1120s, rf.form_1065_llc,
    rf.corp_officer, rf.business_owner, rf.form_990_1041
FROM returns r
JOIN clients c ON c.id = r.client_id
LEFT JOIN payments p ON p.return_id = r.id
LEFT JOIN return_forms rf ON rf.return_id = r.id
"""

# ── DB helpers ────────────────────────────────────────────────────────────────

def _enrich(r: dict) -> dict:
    total = r.get("total_fee") or 0
    paid  = r.get("fee_paid")  or 0
    r["balance"]      = round(total - paid, 2) if total else None
    r["paid_in_full"] = bool(total and paid >= total)
    r["color"]        = STATUS_DOT.get(r.get("client_status") or "", "dot-slate")
    r["badge_class"]  = STATUS_BADGE.get(r.get("client_status") or "", "bg-slate-100 text-slate-500 border-slate-200")
    first = r.get("first_name") or ""
    last  = r.get("last_name")  or ""
    r["name_full"] = r.get("display_name") or (f"{last}, {first}".strip(", ") if first else last)
    r["forms"]        = _form_badges(r)
    intake_dt = _parse_iso_date(r.get("intake_date"))
    completion_dt = _parse_iso_date(r.get("logout_date")) or _parse_iso_date(r.get("ack_date"))
    r["cycle_days"] = (
        (completion_dt - intake_dt).days
        if intake_dt and completion_dt and completion_dt >= intake_dt
        else None
    )
    r["late_intake_flag"] = (
        bool(intake_dt) and
        (intake_dt.month > LATE_INTAKE_MONTH or (intake_dt.month == LATE_INTAKE_MONTH and intake_dt.day >= LATE_INTAKE_DAY))
    )
    r["slow_cycle_flag"] = bool(r["cycle_days"] is not None and r["cycle_days"] >= SLOW_CYCLE_DAYS)
    r["risk_flags"] = []
    if r["late_intake_flag"]:
        r["risk_flags"].append("LATE INTAKE")
    if r["slow_cycle_flag"]:
        r["risk_flags"].append("SLOW CYCLE")
    if privacy_mode_enabled():
        r = _mask_return_payload(r)
    return r


def _form_badges(r: dict) -> list[str]:
    mapping = [
        ("form_1040",    "1040"),
        ("sched_a_d",    "A&D"),
        ("sched_c",      "SCH C"),
        ("sched_e",      "SCH E"),
        ("form_1120",    "1120"),
        ("form_1120s",   "1120S"),
        ("form_1065_llc","1065"),
        ("corp_officer", "CORP"),
        ("business_owner","BUS"),
        ("form_990_1041","990"),
    ]
    badges = [label for field, label in mapping if r.get(field)]
    if r.get("is_amended"):   badges.append("1040X")
    if r.get("has_w7"):       badges.append("W7")
    if r.get("is_extension"): badges.append("EXT")
    if r.get("transfer_flag") or r.get("transfer_2025_flag") or r.get("transfer_2026_flag"):
        badges.append("XFER")
    return badges


def _parse_iso_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def query_returns(filters: dict | None = None) -> list[dict]:
    conn = get_connection()
    f = filters or {}
    clauses: list[str] = []
    params:  list      = []

    # "year" here is the INTAKE/SEASON year (e.g. 2026 = the 2025-2026 filing season).
    # A return belongs to season Y if it was brought in during calendar year Y,
    # OR if it has no intake date but its tax_year = Y-1 (Drake-imported TY2025 records).
    year = f.get("year") or date.today().year
    clauses.append(
        "(strftime('%Y', r.intake_date) = ? OR "
        "(r.intake_date IS NULL AND r.tax_year = ?))"
    )
    params.append(str(year))
    params.append(year - 1)

    if f.get("status"):
        statuses = f["status"] if isinstance(f["status"], list) else [f["status"]]
        statuses = [s for s in statuses if s]
        if statuses:
            clauses.append(f"r.client_status IN ({','.join('?' for _ in statuses)})")
            params.extend(statuses)

    if f.get("processor"):
        clauses.append("r.processor = ?")
        params.append(f["processor"])

    if f.get("balance_due"):
        clauses.append(
            "(p.total_fee IS NOT NULL AND COALESCE(p.fee_paid,0) < p.total_fee)"
        )
    if f.get("late_intake"):
        clauses.append(
            "(r.intake_date IS NOT NULL AND ("
            "CAST(substr(r.intake_date,6,2) AS INTEGER) > 4 OR "
            "(CAST(substr(r.intake_date,6,2) AS INTEGER) = 4 AND CAST(substr(r.intake_date,9,2) AS INTEGER) >= 1)"
            "))"
        )
    if f.get("slow_cycle"):
        clauses.append(
            "(r.intake_date IS NOT NULL AND "
            "(r.logout_date IS NOT NULL OR r.ack_date IS NOT NULL) AND "
            "(julianday(COALESCE(r.logout_date, r.ack_date)) - julianday(r.intake_date)) >= ?)"
        )
        params.append(SLOW_CYCLE_DAYS)

    if f.get("form"):
        form_col = f["form"]
        allowed = {
            "form_1040", "sched_a_d", "sched_c", "sched_e",
            "form_1120", "form_1120s", "form_1065_llc",
            "corp_officer", "business_owner", "form_990_1041",
            "is_amended", "has_w7", "is_extension",
        }
        if form_col in allowed:
            # is_amended/has_w7/is_extension live on returns; forms live on return_forms
            if form_col in ("is_amended", "has_w7", "is_extension"):
                clauses.append(f"r.{form_col} = 1")
            else:
                clauses.append(f"rf.{form_col} = 1")

    if f.get("q"):
        q = f["q"].strip()
        if q.isdigit():
            clauses.append("r.log_number = ?")
            params.append(q)
        else:
            qp = f"%{q.lower()}%"
            clauses.append(
                "(lower(c.last_name) LIKE ? OR lower(c.first_name) LIKE ? OR lower(COALESCE(c.display_name,'')) LIKE ?)"
            )
            params.extend([qp, qp, qp])

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    sql   = f"{_SELECT} {where} ORDER BY CAST(r.log_number AS INTEGER), r.id"

    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [_enrich(dict(r)) for r in rows]


def get_one(return_id: int) -> dict | None:
    conn = get_connection()
    row  = conn.execute(f"{_SELECT} WHERE r.id = ?", (return_id,)).fetchone()
    conn.close()
    return _enrich(dict(row)) if row else None


def get_status_counts(year: int) -> dict[str, int]:
    conn = get_connection()
    rows = conn.execute(
        "SELECT client_status, COUNT(*) n FROM returns "
        "WHERE (strftime('%Y', intake_date) = ? OR (intake_date IS NULL AND tax_year = ?)) "
        "GROUP BY client_status",
        (str(year), year - 1),
    ).fetchall()
    conn.close()
    return {r["client_status"]: r["n"] for r in rows if r["client_status"]}


def get_totals(year: int) -> dict:
    conn = get_connection()
    row = conn.execute(
        """
        SELECT
            COALESCE(SUM(p.total_fee),0) billed,
            COALESCE(SUM(p.fee_paid),0)  collected
        FROM returns r
        LEFT JOIN payments p ON p.return_id = r.id
        WHERE (strftime('%Y', r.intake_date) = ? OR (r.intake_date IS NULL AND r.tax_year = ?))
        """,
        (str(year), year - 1),
    ).fetchone()
    conn.close()
    billed    = row["billed"]    if row else 0
    collected = row["collected"] if row else 0
    return {"billed": billed, "collected": collected, "outstanding": round(billed - collected, 2)}


def get_processors(year: int) -> list[str]:
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT processor FROM returns "
        "WHERE (strftime('%Y', intake_date) = ? OR (intake_date IS NULL AND tax_year = ?)) "
        "AND processor IS NOT NULL ORDER BY processor",
        (str(year), year - 1),
    ).fetchall()
    conn.close()
    return [r["processor"] for r in rows]


def base_ctx(year: int | None = None) -> dict:
    today_year = date.today().year
    # Never let the season picker go backwards to a tax year.
    # Callers should pass the intake/season year, not the tax year.
    y = year if (year and year >= today_year) else today_year
    conn = get_connection()
    pending_review = conn.execute(
        "SELECT COUNT(*) n FROM review_queue WHERE status='pending'"
    ).fetchone()["n"]
    # Rejected returns — always pulled regardless of season filter
    rejected_rows = conn.execute(
        f"{_SELECT} WHERE r.client_status = 'REJECTED' ORDER BY r.updated_at DESC"
    ).fetchall()
    conn.close()
    rejected = [_enrich(dict(r)) for r in rejected_rows]
    return {
        "current_year":         y,
        "status_flow":          STATUS_FLOW,
        "status_badge":         STATUS_BADGE,
        "status_dot":           STATUS_DOT,
        "status_counts":        get_status_counts(y),
        "totals":               get_totals(y),
        "processors":           get_processors(y),
        "app_env":              APP_ENV,
        "privacy_mode":         privacy_mode_enabled(),
        "pending_review_count": pending_review,
        "rejected_returns":     rejected,
        "rejected_count":       len(rejected),
    }


def build_client_habit_profile(conn, client_id: int, target_year: int | None = None) -> dict:
    """
    Build planning reminders from prior-year filing behavior.
    This is intentionally heuristic so staff can anticipate complexity
    while still re-confirming items that tend to change year to year.
    """
    rows = conn.execute(
        """
        SELECT
            r.id, r.tax_year, r.intake_date, r.is_extension, r.has_w7,
            rf.sched_c, rf.sched_e, rf.form_1120, rf.form_1120s,
            rf.form_1065_llc, rf.business_owner, rf.corp_officer,
            r.insurance_type
        FROM returns r
        LEFT JOIN return_forms rf ON rf.return_id = r.id
        WHERE r.client_id = ?
        ORDER BY r.tax_year DESC, r.id DESC
        """,
        (client_id,),
    ).fetchall()

    if not rows:
        return {
            "target_year": target_year or date.today().year,
            "risk_level": "standard",
            "late_filer": False,
            "late_years": [],
            "recurring_forms": [],
            "ask_again": [],
            "reminders": [],
        }

    effective_target_year = target_year or date.today().year
    prior_rows = [r for r in rows if (r["tax_year"] or 0) < effective_target_year] or rows

    def _is_late_intake(intake_date: str | None) -> bool:
        if not intake_date or len(intake_date) < 7:
            return False
        try:
            month = int(intake_date[5:7])
            day = int(intake_date[8:10]) if len(intake_date) >= 10 else 1
            return month >= 4 or (month == 3 and day >= 25)
        except ValueError:
            return False

    late_years = sorted(
        [r["tax_year"] for r in prior_rows if r["tax_year"] and _is_late_intake(r["intake_date"])],
        reverse=True,
    )
    slow_cycle_years: list[int] = []
    for r in prior_rows:
        start = _parse_iso_date(r["intake_date"])
        end = _parse_iso_date(r["logout_date"])
        if start and end and end >= start and (end - start).days >= SLOW_CYCLE_DAYS and r["tax_year"]:
            slow_cycle_years.append(r["tax_year"])
    slow_cycle_years = sorted(set(slow_cycle_years), reverse=True)

    recurring_forms: list[str] = []
    form_rules = [
        ("sched_c", "Schedule C"),
        ("sched_e", "Schedule E"),
        ("form_1065_llc", "K-1/1065 partnership"),
        ("form_1120", "1120 corporate"),
        ("form_1120s", "1120S"),
        ("business_owner", "Business owner"),
        ("corp_officer", "Corporate officer"),
        ("is_extension", "Extension filing"),
        ("has_w7", "W-7 / ITIN"),
    ]
    for key, label in form_rules:
        if any(r[key] for r in prior_rows):
            recurring_forms.append(label)

    ask_again: list[str] = []
    had_marketplace = any(
        (r["insurance_type"] or "").strip().lower().startswith("marketplace")
        for r in prior_rows
    )
    if had_marketplace:
        ask_again.append("Confirm current 1095-A / Marketplace coverage for this year")
    if any((r["insurance_type"] or "").strip().lower() in {"medi-cal", "medicare"} for r in prior_rows):
        ask_again.append("Reconfirm current Medi-Cal/Medicare status (can change year to year)")

    reminders: list[str] = []
    if late_years:
        years = ", ".join(str(y) for y in late_years[:3])
        reminders.append(
            f"Historically filed late ({years}). Trigger early outreach before March."
        )
    if recurring_forms:
        reminders.append(
            "Prior complexity detected: " + ", ".join(recurring_forms[:5]) +
            (", ..." if len(recurring_forms) > 5 else "")
        )
    if slow_cycle_years:
        years = ", ".join(str(y) for y in slow_cycle_years[:3])
        reminders.append(
            f"Historically long turnaround ({years}). Ask for missing docs at intake to avoid delays."
        )
    reminders.extend(ask_again)

    risk_level = "high" if (len(late_years) >= 2 or len(recurring_forms) >= 3 or len(slow_cycle_years) >= 2) else "watch"
    if not late_years and len(recurring_forms) <= 1:
        risk_level = "standard"

    return {
        "target_year": effective_target_year,
        "risk_level": risk_level,
        "late_filer": bool(late_years),
        "late_years": late_years,
        "slow_cycle_years": slow_cycle_years,
        "recurring_forms": recurring_forms,
        "ask_again": ask_again,
        "reminders": reminders,
    }


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if username == _LOGIN_USER and password == _LOGIN_PASS:
            session["logged_in"] = True
            session["username"]  = username
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Page routes ───────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    year = int(request.args.get("year", date.today().year))
    filters = {
        "year":        year,
        "status":      request.args.getlist("status") or None,
        "processor":   request.args.get("processor"),
        "balance_due": request.args.get("balance_due"),
        "late_intake": request.args.get("late_intake"),
        "slow_cycle":  request.args.get("slow_cycle"),
        "form":        request.args.get("form"),
        "q":           request.args.get("q"),
    }
    returns = query_returns(filters)
    ctx = base_ctx(year)
    ctx.update({"active_page": "dashboard", "returns": returns, "filters": filters})
    return render_template("dashboard.html", **ctx)


@app.route("/return/<int:return_id>")
@login_required
def return_detail(return_id: int):
    ret = get_one(return_id)
    if not ret:
        abort(404)
    conn   = get_connection()
    notes  = conn.execute(
        "SELECT * FROM notes WHERE return_id=? ORDER BY created_at DESC", (return_id,)
    ).fetchall()
    events = conn.execute(
        "SELECT * FROM status_events WHERE return_id=? ORDER BY event_timestamp DESC", (return_id,)
    ).fetchall()
    conn.close()
    notes_payload = [dict(n) for n in notes]
    if privacy_mode_enabled():
        for note in notes_payload:
            note["note_text"] = _mask_value(note.get("note_text"))

    # Always use current calendar year for the season picker — never the return's tax year.
    ctx = base_ctx(date.today().year)
    ctx.update({
        "active_page": "dashboard",
        "ret":    ret,
        "notes":  notes_payload,
        "events": [dict(e) for e in events],
    })
    return render_template("return_detail.html", **ctx)


@app.route("/logout-queue")
@login_required
def logout_queue():
    year = int(request.args.get("year", date.today().year))
    conn = get_connection()
    rows = conn.execute(
        f"{_SELECT} WHERE (strftime('%Y', r.intake_date) = ? OR (r.intake_date IS NULL AND r.tax_year = ?)) "
        "AND r.client_status = 'PICKUP' ORDER BY CAST(r.log_number AS INTEGER)",
        (str(year), year - 1),
    ).fetchall()
    conn.close()
    ctx = base_ctx(year)
    ctx.update({
        "active_page": "logout",
        "returns":     [_enrich(dict(r)) for r in rows],
        "today":       date.today().isoformat(),
    })
    return render_template("logout_queue.html", **ctx)


@app.route("/payments")
@login_required
def payments():
    year         = int(request.args.get("year", date.today().year))
    balance_only = request.args.get("balance_only")
    where = "WHERE r.tax_year=?"
    if balance_only:
        where += " AND p.total_fee IS NOT NULL AND COALESCE(p.fee_paid,0) < p.total_fee"
    conn = get_connection()
    rows = conn.execute(
        f"{_SELECT} {where} ORDER BY CAST(r.log_number AS INTEGER)", (year,)
    ).fetchall()
    conn.close()
    ctx = base_ctx(year)
    ctx.update({
        "active_page":  "payments",
        "returns":      [_enrich(dict(r)) for r in rows],
        "balance_only": balance_only,
    })
    return render_template("payments.html", **ctx)


# ── Intake form ───────────────────────────────────────────────────────────────

@app.route("/intake", methods=["GET", "POST"])
@login_required
def intake():
    if request.method == "GET":
        ctx = base_ctx()
        ctx.update({
            "active_page": "intake",
            "today": date.today().isoformat(),
            "error": None,
            "habit_profile": None,
        })
        return render_template("intake.html", **ctx)

    # POST — create records
    f = request.form
    ts = now()
    today_iso = date.today().isoformat()

    last_name  = (f.get("last_name") or "").strip().upper()
    first_name = (f.get("first_name") or "").strip().upper()
    if not last_name:
        ctx = base_ctx()
        ctx.update({"active_page": "intake", "today": today_iso, "error": "Last name is required.", "prefill": {}})
        return render_template("intake.html", **ctx), 400

    def _v(key):
        val = f.get(key, "").strip()
        return val or None

    def _n(key):
        val = f.get(key, "").strip()
        try:
            return float(val) if val else None
        except ValueError:
            return None

    def _i(key):
        val = f.get(key, "").strip()
        return int(val) if val.isdigit() else None

    conn = get_connection()
    try:
        tax_year = _i("tax_year") or date.today().year

        # ── Auto log number (max + 1 for this tax year) ───────────────────────
        row = conn.execute(
            "SELECT MAX(CAST(log_number AS INTEGER)) AS mx FROM returns WHERE tax_year = ?",
            (tax_year,),
        ).fetchone()
        log_number = str((row["mx"] or 0) + 1)

        # ── Client — insert new or update existing (re-intake) ────────────────
        existing_client_id = _i("client_id")
        if existing_client_id:
            conn.execute(
                """
                UPDATE clients SET
                    last_name=?, first_name=?, ssn_last4=?,
                    spouse_last_name=?, spouse_first_name=?,
                    taxpayer_dob=?, spouse_dob=?,
                    taxpayer_occupation=?, spouse_occupation=?,
                    taxpayer_phone=?, taxpayer_cell=?, taxpayer_work_phone=?,
                    spouse_cell=?, spouse_work_phone=?,
                    taxpayer_email=?, spouse_email=?,
                    address=?, referral_flag=?, referred_by=?,
                    is_new_client=0, prior_year_log=?, updated_at=?
                WHERE id=?
                """,
                (
                    last_name, first_name, _v("ssn_last4"),
                    (_v("spouse_last_name") or "").upper() or None,
                    (_v("spouse_first_name") or "").upper() or None,
                    _v("taxpayer_dob"), _v("spouse_dob"),
                    _v("taxpayer_occupation"), _v("spouse_occupation"),
                    _v("taxpayer_phone"), _v("taxpayer_cell"), _v("taxpayer_work_phone"),
                    _v("spouse_cell"), _v("spouse_work_phone"),
                    _v("taxpayer_email"), _v("spouse_email"),
                    _v("address"),
                    1 if f.get("referral_flag") else 0,
                    _v("referred_by"),
                    _v("prior_year_log"),
                    ts, existing_client_id,
                ),
            )
            client_id = existing_client_id
        else:
            conn.execute(
                """
                INSERT INTO clients (
                    last_name, first_name, ssn_last4,
                    spouse_last_name, spouse_first_name,
                    taxpayer_dob, spouse_dob,
                    taxpayer_occupation, spouse_occupation,
                    taxpayer_phone, taxpayer_cell, taxpayer_work_phone,
                    spouse_cell, spouse_work_phone,
                    taxpayer_email, spouse_email,
                    address, referral_flag, referred_by,
                    is_new_client, prior_year_log,
                    created_at, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    last_name, first_name, _v("ssn_last4"),
                    (_v("spouse_last_name") or "").upper() or None,
                    (_v("spouse_first_name") or "").upper() or None,
                    _v("taxpayer_dob"), _v("spouse_dob"),
                    _v("taxpayer_occupation"), _v("spouse_occupation"),
                    _v("taxpayer_phone"), _v("taxpayer_cell"), _v("taxpayer_work_phone"),
                    _v("spouse_cell"), _v("spouse_work_phone"),
                    _v("taxpayer_email"), _v("spouse_email"),
                    _v("address"),
                    1 if f.get("referral_flag") else 0,
                    _v("referred_by"),
                    int(f.get("is_new_client", "0")),
                    _v("prior_year_log"),
                    ts, ts,
                ),
            )
            client_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ── Return ────────────────────────────────────────────────────────────
        conn.execute(
            """
            INSERT INTO returns (
                client_id, log_number, tax_year, client_status,
                processor, verified, intake_date, interview_by,
                filing_status, promise_date, delivered_by,
                date_signatures_emailed, date_reports_emailed,
                overtime_flag, insurance_type, digital_assets,
                bank_name, bank_routing, bank_account, bank_account_type,
                notes_intake,
                is_amended, has_w7, is_extension,
                estimate_irs, estimate_state, final_irs, final_state,
                created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                client_id,
                log_number,
                tax_year,
                "PROCESSING",
                _v("processor"),
                1 if f.get("verified") else 0,
                _v("intake_date") or today_iso,
                _v("interview_by"),
                _v("filing_status"),
                _v("promise_date"),
                _v("delivered_by"),
                _v("date_signatures_emailed"),
                _v("date_reports_emailed"),
                int(f.get("overtime_flag", "0")),
                _v("insurance_type"),
                int(f.get("digital_assets", "0")),
                _v("bank_name"), _v("bank_routing"), _v("bank_account"), _v("bank_account_type"),
                _v("notes_intake"),
                1 if f.get("is_amended") else None,
                1 if f.get("has_w7") else None,
                1 if f.get("is_extension") else None,
                _n("estimate_irs"), _n("estimate_state"),
                _n("final_irs"), _n("final_state"),
                ts, ts,
            ),
        )
        return_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ── Return forms ──────────────────────────────────────────────────────
        form_fields = [
            "form_1040", "sched_a_d", "sched_c", "sched_e",
            "form_1120", "form_1120s", "form_1065_llc",
            "corp_officer", "business_owner", "form_990_1041",
        ]
        form_vals = {field: (1 if f.get(field) else None) for field in form_fields}
        conn.execute(
            f"""INSERT INTO return_forms (return_id, {', '.join(form_fields)})
                VALUES (?, {', '.join('?' for _ in form_fields)})""",
            [return_id] + [form_vals[k] for k in form_fields],
        )

        # ── Payment ───────────────────────────────────────────────────────────
        conn.execute(
            """
            INSERT INTO payments (
                return_id, total_fee, fee_paid, receipt_number, receipt2_number,
                accounting_fee, w7_fee, form_1099_fee, license_fee,
                reprocess_fee, discount_amount, special_discount, down_payment
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                return_id,
                _n("total_fee"), _n("fee_paid"),
                _v("receipt_number"), _v("receipt2_number"),
                _n("accounting_fee"), _n("w7_fee"), _n("form_1099_fee"), _n("license_fee"),
                _n("reprocess_fee"), _n("discount_amount"), _n("special_discount"),
                _n("down_payment"),
            ),
        )

        # ── Dependents ────────────────────────────────────────────────────────
        dep_count = int(f.get("dep_count", "6"))
        for i in range(1, dep_count + 1):
            name = (f.get(f"dep_name_{i}") or "").strip().upper()
            if not name:
                continue
            conn.execute(
                """
                INSERT INTO dependents
                  (return_id, full_name, ssn_last4, relationship, date_of_birth, medi_cal, created_at)
                VALUES (?,?,?,?,?,?,?)
                """,
                (
                    return_id, name,
                    _v(f"dep_ssn_{i}"),
                    _v(f"dep_rel_{i}"),
                    _v(f"dep_dob_{i}"),
                    1 if f.get(f"dep_medicaid_{i}") else 0,
                    ts,
                ),
            )

        # ── Status event ──────────────────────────────────────────────────────
        conn.execute(
            """
            INSERT INTO status_events
              (return_id, event_type, old_status, new_status, event_timestamp, source_file, note)
            VALUES (?, 'STATUS_CHANGED', NULL, 'PROCESSING', ?, 'INTAKE', 'Created via intake form')
            """,
            (return_id, ts),
        )

        # ── Notes ─────────────────────────────────────────────────────────────
        if _v("notes_intake"):
            conn.execute(
                "INSERT INTO notes (return_id, note_text, source, created_at) VALUES (?,?,'INTAKE',?)",
                (return_id, _v("notes_intake"), ts),
            )

        conn.commit()
        return redirect(f"/return/{return_id}")

    except Exception as exc:
        conn.rollback()
        ctx = base_ctx()
        ctx.update({"active_page": "intake", "today": today_iso, "error": str(exc)})
        return render_template("intake.html", **ctx), 500
    finally:
        conn.close()


# ── CSV Upload / Analyze ──────────────────────────────────────────────────────

@app.route("/upload", methods=["GET"])
@login_required
def upload_get():
    ctx = base_ctx()
    ctx.update({"active_page": "upload", "error": None})
    return render_template("upload.html", **ctx)


@app.route("/upload/preview", methods=["POST"])
@login_required
def upload_preview():
    f = request.files.get("csv_file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded"}), 400

    file_bytes = f.read()
    result = analyze(file_bytes, f.filename)

    # Store file bytes in a temp file keyed by a token for the confirm step
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", prefix="taxops_upload_")
    tmp.write(file_bytes)
    tmp.close()

    cols = [
        {
            "index":      c.col_index,
            "raw_header": c.raw_header,
            "table":      c.table,
            "field":      c.field,
            "field_type": c.field_type,
            "confidence": c.confidence,
            "skip":       c.skip,
        }
        for c in result.columns
    ]

    return jsonify({
        "tmp_path":    tmp.name,
        "filename":    f.filename,
        "total_rows":  result.total_rows,
        "warnings":    result.warnings,
        "columns":     cols,
        "sample_rows": result.sample_rows[:8],
        "data_start":  result.data_start_index,
        "header_row":  result.header_row_index,
    })


@app.route("/upload/confirm", methods=["POST"])
@login_required
def upload_confirm():
    """Execute import using the analysis result confirmed by staff."""
    data       = request.get_json(force=True)
    tmp_path   = data.get("tmp_path", "")
    overrides  = data.get("overrides", {})   # {str(col_index): "table.field" | "skip"}
    tax_year   = int(data.get("tax_year", date.today().year))

    if not tmp_path or not os.path.exists(tmp_path):
        return jsonify({"error": "Upload session expired — please re-upload."}), 400

    with open(tmp_path, "rb") as fh:
        file_bytes = fh.read()

    result    = analyze(file_bytes)
    ts        = now()
    today_iso = date.today().isoformat()

    # Apply overrides to column mappings
    for c in result.columns:
        key = str(c.col_index)
        if key in overrides:
            ov = overrides[key]
            if ov == "skip":
                c.skip = True
                c.table = c.field = ""
            elif "." in ov:
                parts = ov.split(".", 1)
                c.table, c.field = parts[0], parts[1]
                c.skip = False

    rows = iter_data_rows(file_bytes, result)

    conn = get_connection()
    stats = {"created": 0, "updated": 0, "skipped": 0, "review": 0, "errors": []}

    try:
        # Build client cache once so fuzzy matching doesn't hammer the DB per row
        client_cache = _all_clients_cache(conn)
        for row_data in rows:
            try:
                _import_row(conn, row_data, tax_year, ts, today_iso, stats, _client_cache=client_cache)
            except Exception as exc:
                stats["errors"].append(str(exc))
                if len(stats["errors"]) > 20:
                    break
        conn.commit()
    finally:
        conn.close()
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    return jsonify(stats)


def _import_row_forced(conn, row_data: dict, tax_year: int, ts: str, today_iso: str,
                       stats: dict, *, client_id: int | None):
    """Run _import_row but skip fuzzy matching — use the supplied client_id directly,
    or create a new client if client_id is None."""
    # Temporarily patch the row so _import_row's name-parse produces something
    # that will definitely match (or not) based on client_id override.
    # Easiest: delegate to _import_row with a single-entry cache that forces the match.
    if client_id is not None:
        forced_cache = [{"id": client_id, "ln": "\x00FORCED\x00", "fn": ""}]
        # Pre-seed row name with the sentinel so exact match fires
        patched = dict(row_data)
        patched["clients.last_name"]  = "\x00FORCED\x00"
        patched["clients.first_name"] = ""
        _import_row(conn, patched, tax_year, ts, today_iso, stats, _client_cache=forced_cache)
    else:
        # No client_id → force new client by using an empty cache
        _import_row(conn, row_data, tax_year, ts, today_iso, stats, _client_cache=[])


def _import_row(conn, row_data: dict, tax_year: int, ts: str, today_iso: str, stats: dict,
                _client_cache: list | None = None):
    """Import a single analyzed row into the database."""
    def g(table, field):
        return row_data.get(f"{table}.{field}", "") or ""

    raw_last  = normalize_string(g("clients", "last_name"))
    raw_first = normalize_string(g("clients", "first_name"))
    display   = normalize_string(g("clients", "display_name"))

    if not raw_last and display:
        raw_last = display

    log_number = normalize_string(g("returns", "log_number"))

    # Placeholder row: reserved log slot with no client data — skip silently
    if not raw_last:
        stats["skipped"] += 1
        return

    if not log_number:
        stats["skipped"] += 1
        return

    # ── Parse name via name_matcher ───────────────────────────────────────────
    last_name, first_name = parse_name(raw_last)
    # If the CSV already split first/last, prefer that
    if raw_first:
        first_name = raw_first.upper().strip() or None
    last_name = last_name.upper().strip()
    first_name = (first_name or "").upper().strip() or None

    # ── Match or create client ────────────────────────────────────────────────
    match = fuzzy_find_client(conn, last_name, first_name, cache=_client_cache)

    if match and not match["needs_review"]:
        # Confident match — upsert against existing client
        client_id = match["client_id"]
        conn.execute("UPDATE clients SET updated_at=? WHERE id=?", (ts, client_id))
        stats["updated"] = stats.get("updated", 0) + 1
        match_method = match["method"]
    elif match and match["needs_review"]:
        # Low-confidence — park in review queue for human decision, don't process yet
        stats["review"] = stats.get("review", 0) + 1
        raw_yr = g("returns", "tax_year")
        ret_year_q = int(raw_yr) if raw_yr.isdigit() else tax_year
        conn.execute(
            """INSERT INTO review_queue
               (status, csv_last, csv_first, csv_log, csv_year,
                proposed_client_id, match_score, match_method,
                raw_json, reason, created_at)
               VALUES ('pending',?,?,?,?,?,?,?,?,?,?)""",
            (
                last_name, first_name, log_number, ret_year_q,
                match["client_id"], match["score"], match["method"],
                json.dumps(row_data),
                f"Fuzzy score={match['score']} method={match['method']}",
                ts,
            ),
        )
        return  # do not create return — wait for staff to resolve
    else:
        # No match — create new client
        conn.execute(
            """INSERT INTO clients (last_name, first_name, referral_flag, referred_by,
                                    prior_year_log, is_new_client, created_at, updated_at)
               VALUES (?,?,?,?,?,1,?,?)""",
            (
                last_name, first_name,
                1 if g("clients", "referral_flag") else 0,
                normalize_string(g("clients", "referred_by")),
                normalize_string(g("clients", "prior_year_log")),
                ts, ts,
            ),
        )
        client_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Add to cache so subsequent rows for the same new client match
        if _client_cache is not None:
            _client_cache.append({"id": client_id, "ln": last_name.upper(), "fn": (first_name or "").upper()})
        stats["created"] = stats.get("created", 0) + 1
        match_method = "new"

    # ── Match or create return ────────────────────────────────────────────────
    ret_year = tax_year
    raw_yr = g("returns", "tax_year")
    if raw_yr.isdigit() and 2000 <= int(raw_yr) <= 2030:
        ret_year = int(raw_yr)

    # Match by client + year only — log_number may be absent on Drake-imported
    # returns and will be written onto the record if the CSV supplies it.
    existing_ret = conn.execute(
        "SELECT id, log_number FROM returns WHERE client_id=? AND tax_year=?",
        (client_id, ret_year),
    ).fetchone()

    raw_status  = g("returns", "client_status")
    norm_status = normalize_status(raw_status) if raw_status else "PROCESSING"

    intake_dt,  _ = normalize_date(g("returns", "intake_date"))
    pickup_dt,  _ = normalize_date(g("returns", "pickup_date"))
    logout_dt,  _ = normalize_date(g("returns", "logout_date"))
    emailed_dt, _ = normalize_date(g("returns", "date_emailed"))
    updated_dt, _ = normalize_date(g("returns", "updated_date"))

    def flag(table, field):
        v = g(table, field).strip()
        return 1 if v and v not in ("0", "", " ") else None

    ret_fields = dict(
        client_id=client_id,
        log_number=log_number,
        tax_year=ret_year,
        client_status=norm_status,
        processor=normalize_string(g("returns", "processor")),
        verified=flag("returns", "verified"),
        intake_date=intake_dt or today_iso,
        date_emailed=emailed_dt,
        pickup_date=pickup_dt,
        logout_date=logout_dt,
        updated_date=updated_dt,
        is_amended=flag("returns", "is_amended"),
        has_w7=flag("returns", "has_w7"),
        is_extension=flag("returns", "is_extension"),
        transfer_flag=flag("returns", "transfer_flag"),
        transfer_2025_flag=flag("returns", "transfer_2025_flag"),
        transfer_2026_flag=flag("returns", "transfer_2026_flag"),
    )

    if existing_ret:
        ret_id = existing_ret["id"]
        # Stamp log_number from CSV onto the return if it didn't have one yet
        # (Drake imports don't carry log numbers; the manual log is the source).
        existing_log = existing_ret["log_number"]
        new_log = log_number if log_number else existing_log
        conn.execute(
            """UPDATE returns
               SET log_number=?,
                   client_status=?, processor=?, verified=?,
                   intake_date=COALESCE(intake_date,?),
                   pickup_date=COALESCE(pickup_date,?),
                   logout_date=COALESCE(logout_date,?),
                   updated_at=?
               WHERE id=?""",
            (new_log, norm_status, ret_fields["processor"], ret_fields["verified"],
             ret_fields["intake_date"], pickup_dt, logout_dt, ts, ret_id),
        )
    else:
        cols = ", ".join(ret_fields.keys()) + ", created_at, updated_at"
        vals = ", ".join("?" for _ in ret_fields) + ", ?, ?"
        conn.execute(
            f"INSERT INTO returns ({cols}) VALUES ({vals})",
            list(ret_fields.values()) + [ts, ts],
        )
        ret_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # ── Forms ─────────────────────────────────────────────────────────────────
    form_fields = {
        "form_1040": flag("return_forms", "form_1040"),
        "sched_a_d": flag("return_forms", "sched_a_d"),
        "sched_c":   flag("return_forms", "sched_c"),
        "sched_e":   flag("return_forms", "sched_e"),
        "form_1120": flag("return_forms", "form_1120"),
        "form_1120s":flag("return_forms", "form_1120s"),
        "form_1065_llc": flag("return_forms", "form_1065_llc"),
        "corp_officer":  flag("return_forms", "corp_officer"),
        "business_owner":flag("return_forms", "business_owner"),
        "form_990_1041": flag("return_forms", "form_990_1041"),
    }
    existing_forms = conn.execute("SELECT id FROM return_forms WHERE return_id=?", (ret_id,)).fetchone()
    if existing_forms:
        sets = ", ".join(f"{k}=?" for k in form_fields)
        conn.execute(f"UPDATE return_forms SET {sets} WHERE return_id=?",
                     list(form_fields.values()) + [ret_id])
    else:
        fcols = "return_id, " + ", ".join(form_fields.keys())
        fvals = "?, " + ", ".join("?" for _ in form_fields)
        conn.execute(f"INSERT INTO return_forms ({fcols}) VALUES ({fvals})",
                     [ret_id] + list(form_fields.values()))

    # ── Payment ───────────────────────────────────────────────────────────────
    total_fee = normalize_currency(g("payments", "total_fee"))
    fee_paid  = normalize_currency(g("payments", "fee_paid"))
    cc_fee    = normalize_currency(g("payments", "cc_fee"))
    receipt   = normalize_string(g("payments", "receipt_number"))
    zelle     = normalize_string(g("payments", "zelle_or_check_ref"))
    cash      = normalize_string(g("payments", "cash_or_qpay_ref"))

    if any(v is not None for v in [total_fee, fee_paid, cc_fee, receipt]):
        existing_pay = conn.execute("SELECT id FROM payments WHERE return_id=?", (ret_id,)).fetchone()
        if existing_pay:
            conn.execute(
                """UPDATE payments SET
                   total_fee=COALESCE(?,total_fee), fee_paid=COALESCE(?,fee_paid),
                   cc_fee=COALESCE(?,cc_fee), receipt_number=COALESCE(?,receipt_number),
                   zelle_or_check_ref=COALESCE(?,zelle_or_check_ref),
                   cash_or_qpay_ref=COALESCE(?,cash_or_qpay_ref)
                   WHERE return_id=?""",
                (total_fee, fee_paid, cc_fee, receipt, zelle, cash, ret_id),
            )
        else:
            conn.execute(
                """INSERT INTO payments
                   (return_id, total_fee, fee_paid, cc_fee, receipt_number,
                    zelle_or_check_ref, cash_or_qpay_ref)
                   VALUES (?,?,?,?,?,?,?)""",
                (ret_id, total_fee, fee_paid, cc_fee, receipt, zelle, cash),
            )

    # ── Note ──────────────────────────────────────────────────────────────────
    note_text = normalize_string(g("notes", "note_text"))
    if note_text:
        dup = conn.execute(
            "SELECT id FROM notes WHERE return_id=? AND lower(note_text)=lower(?)",
            (ret_id, note_text),
        ).fetchone()
        if not dup:
            conn.execute(
                "INSERT INTO notes (return_id, note_text, source, created_at) VALUES (?,?,'CSV_UPLOAD',?)",
                (ret_id, note_text, ts),
            )


# ── Export ────────────────────────────────────────────────────────────────────

@app.route("/export")
@login_required
def export_excel():
    """Export the current filtered view as an .xlsx file."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year = int(request.args.get("year", date.today().year))
    filters = {
        "year":        year,
        "status":      request.args.getlist("status") or None,
        "processor":   request.args.get("processor"),
        "balance_due": request.args.get("balance_due"),
        "late_intake": request.args.get("late_intake"),
        "slow_cycle":  request.args.get("slow_cycle"),
        "form":        request.args.get("form"),
        "q":           request.args.get("q"),
    }
    rows = query_returns(filters)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TaxOps {year}"

    # ── Styles ────────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    DATA_FONT    = Font(name="Calibri", size=11)
    BOLD_FONT    = Font(name="Calibri", bold=True, size=11)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    MONEY        = '#,##0.00'
    thin         = Side(style="thin", color="E2E8F0")
    BORDER       = Border(bottom=thin)

    STATUS_COLORS = {
        "PROCESSING":  "E0F2FE",
        "FINALIZE":    "FEF9C3", "PICKUP":      "CCFBF1",
        "EFILE READY": "E0E7FF", "EFILE":       "EDE9FE",
        "LOG OUT":     "F1F5F9",
    }

    # ── Header row ────────────────────────────────────────────────────────────
    COLUMNS = [
        ("Log #",        9),  ("Last Name",    22), ("First Name",   18),
        ("Year",         7),  ("Status",       14), ("Preparer",     12),
        ("Forms",        18), ("Intake Date",  13), ("Pickup Date",  13),
        ("Logout Date",  13), ("Total Fee",    12), ("Fee Paid",     12),
        ("Balance",      12), ("Receipt #",    13), ("✓",             5),
    ]
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, r in enumerate(rows, start=2):
        status  = r.get("client_status") or ""
        fill_hex = STATUS_COLORS.get(status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_hex)

        forms_str = "  ".join(r.get("forms") or [])
        balance   = r.get("balance") or 0
        total_fee = r.get("total_fee") or 0
        fee_paid  = r.get("fee_paid") or 0

        values = [
            r.get("log_number") or "",
            r.get("last_name")  or "",
            r.get("first_name") or "",
            r.get("tax_year")   or "",
            status,
            r.get("processor")  or "",
            forms_str,
            r.get("intake_date")  or "",
            r.get("pickup_date")  or "",
            r.get("logout_date")  or "",
            total_fee,
            fee_paid,
            balance,
            r.get("receipt_number") or "",
            "✓" if r.get("verified") else "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = BORDER
            cell.font   = DATA_FONT
            # Money columns
            if col_idx in (11, 12, 13) and isinstance(value, (int, float)) and value:
                cell.number_format = MONEY
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                if col_idx == 13 and balance > 0:
                    cell.font = Font(name="Calibri", size=11, bold=True, color="DC2626")
            elif col_idx == 1:
                cell.font      = Font(name="Calibri", bold=True, size=11)
                cell.alignment = CENTER
            elif col_idx in (4, 15):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

        ws.row_dimensions[row_idx].height = 18

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Footer summary ────────────────────────────────────────────────────────
    footer_row = len(rows) + 2
    ws.cell(row=footer_row, column=10, value="TOTALS").font = BOLD_FONT
    total_fee_sum = sum(r.get("total_fee") or 0 for r in rows)
    fee_paid_sum  = sum(r.get("fee_paid")  or 0 for r in rows)
    balance_sum   = sum(r.get("balance")   or 0 for r in rows)
    for col_idx, val in [(11, total_fee_sum), (12, fee_paid_sum), (13, balance_sum)]:
        c = ws.cell(row=footer_row, column=col_idx, value=val)
        c.font         = BOLD_FONT
        c.number_format = MONEY
        c.alignment    = Alignment(horizontal="right", vertical="center")
        if col_idx == 13 and val > 0:
            c.font = Font(name="Calibri", bold=True, size=11, color="DC2626")

    # ── Stream to response ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    status_label = (filters["status"][0] if filters["status"] else "ALL").replace(" ", "-")
    filename = f"TaxOps_{year}_{status_label}.xlsx"

    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Intake log (chronological register) ───────────────────────────────────────

@app.route("/review")
@login_required
def review_queue_page():
    conn = get_connection()
    items = conn.execute(
        """SELECT rq.*,
                  c.last_name  AS db_last,
                  c.first_name AS db_first
           FROM review_queue rq
           LEFT JOIN clients c ON c.id = rq.proposed_client_id
           WHERE rq.status = 'pending'
           ORDER BY rq.id ASC"""
    ).fetchall()
    conn.close()
    ctx = base_ctx()
    ctx.update({"active_page": "review", "items": [dict(i) for i in items]})
    return render_template("review_queue.html", **ctx)


@app.route("/review/resolve", methods=["POST"])
@login_required
def review_resolve():
    """Staff resolves a review_queue item.

    JSON body:
      queue_id  : int
      action    : 'confirm' | 'new' | 'link'
      client_id : int  (required for 'link'; ignored otherwise)
    """
    data     = request.get_json(force=True)
    queue_id = int(data.get("queue_id", 0))
    action   = data.get("action", "")   # confirm | new | link
    override_client_id = data.get("client_id")  # for 'link'

    conn = get_connection()
    item = conn.execute(
        "SELECT * FROM review_queue WHERE id=? AND status='pending'", (queue_id,)
    ).fetchone()

    if not item:
        conn.close()
        return jsonify({"error": "Item not found or already resolved"}), 404

    row_data  = json.loads(item["raw_json"])
    ts        = now()
    today_iso = date.today().isoformat()

    try:
        if action == "new":
            # Force-create a brand-new client by wiping the cache entry
            forced_cache: list = []
            stats = {"created": 0, "updated": 0, "skipped": 0, "review": 0, "errors": []}
            _import_row_forced(conn, row_data, item["csv_year"] or date.today().year,
                               ts, today_iso, stats, client_id=None)

        elif action in ("confirm", "link"):
            cid = override_client_id if action == "link" else item["proposed_client_id"]
            stats = {"created": 0, "updated": 0, "skipped": 0, "review": 0, "errors": []}
            _import_row_forced(conn, row_data, item["csv_year"] or date.today().year,
                               ts, today_iso, stats, client_id=int(cid))
        else:
            conn.close()
            return jsonify({"error": f"Unknown action '{action}'"}), 400

        # Mark resolved
        resolved_cid = override_client_id if action == "link" else (
            item["proposed_client_id"] if action == "confirm" else None
        )
        conn.execute(
            "UPDATE review_queue SET status=?, resolved_client_id=?, resolved_at=? WHERE id=?",
            (action, resolved_cid, ts, queue_id),
        )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(exc)}), 500

    # Return remaining pending count
    remaining = conn.execute(
        "SELECT COUNT(*) n FROM review_queue WHERE status='pending'"
    ).fetchone()["n"]
    conn.close()
    return jsonify({"ok": True, "remaining": remaining, "stats": stats})


@app.route("/intake-log")
@login_required
def intake_log():
    year = int(request.args.get("year", date.today().year))
    conn = get_connection()
    rows = conn.execute(
        f"""
        {_SELECT}
        WHERE (strftime('%Y', r.intake_date) = ? OR (r.intake_date IS NULL AND r.tax_year = ?))
        ORDER BY CAST(r.log_number AS INTEGER) ASC
        """,
        (str(year), year - 1),
    ).fetchall()
    conn.close()

    enriched = [_enrich(dict(r)) for r in rows]

    # Group by intake date
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    for r in enriched:
        key = r.get("intake_date") or "No Date"
        groups[key].append(r)

    sorted_groups = sorted(groups.items(), key=lambda x: x[0])

    ctx = base_ctx(year)
    ctx.update({
        "active_page":  "intake_log",
        "groups":       sorted_groups,
        "total":        len(enriched),
    })
    return render_template("intake_log.html", **ctx)


# ── JSON API ──────────────────────────────────────────────────────────────────

@app.get("/api/clients/search")
@login_required
def api_client_search():
    """Search existing clients by name for re-intake prefill."""
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify([])
    qp = f"%{q.lower()}%"
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT c.id, c.last_name, c.first_name, c.display_name,
               MAX(r.tax_year) AS last_year
        FROM clients c
        LEFT JOIN returns r ON r.client_id = c.id
        WHERE lower(c.last_name) LIKE ? OR lower(c.first_name) LIKE ?
           OR lower(COALESCE(c.display_name,'')) LIKE ?
        GROUP BY c.id
        ORDER BY c.last_name, c.first_name
        LIMIT 12
        """,
        (qp, qp, qp),
    ).fetchall()
    conn.close()
    results = []
    for r in rows:
        first = r["first_name"] or ""
        last  = r["last_name"]  or ""
        name  = r["display_name"] or (f"{last}, {first}".strip(", ") if first else last)
        if privacy_mode_enabled():
            name = f"XXXXX #{r['id']}"
        results.append({
            "id":        r["id"],
            "name":      name,
            "last_year": r["last_year"],
        })
    return jsonify(results)


@app.get("/api/clients/<int:client_id>/reintake")
@login_required
def api_client_reintake(client_id: int):
    """
    Return everything needed to pre-populate the re-intake form for a
    returning client: client fields + most recent return's data + dependents.
    SSN fields are intentionally excluded.
    """
    conn = get_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    if not client:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    # Most recent return for this client
    ret = conn.execute(
        """
        SELECT r.*, rf.form_1040, rf.sched_a_d, rf.sched_c, rf.sched_e,
               rf.form_1120, rf.form_1120s, rf.form_1065_llc,
               rf.corp_officer, rf.business_owner, rf.form_990_1041
        FROM returns r
        LEFT JOIN return_forms rf ON rf.return_id = r.id
        WHERE r.client_id = ?
        ORDER BY r.tax_year DESC, r.id DESC
        LIMIT 1
        """,
        (client_id,),
    ).fetchone()

    # Dependents from that return
    deps = []
    if ret:
        deps = [dict(d) for d in conn.execute(
            "SELECT * FROM dependents WHERE return_id = ? ORDER BY id",
            (ret["id"],),
        ).fetchall()]
        # strip SSN from dependents too
        for d in deps:
            d.pop("ssn_last4", None)
            if privacy_mode_enabled():
                d["full_name"] = _mask_value(d.get("full_name"))
                d["relationship"] = _mask_value(d.get("relationship"))

    habit_profile = build_client_habit_profile(conn, client_id)
    conn.close()

    data = dict(client)
    data.pop("ssn_last4", None)
    if privacy_mode_enabled():
        data = _mask_client_payload(data)

    last_return = {}
    if ret:
        last_return = dict(ret)
        last_return.pop("ssn_last4", None)
        if privacy_mode_enabled():
            last_return = _mask_return_payload(last_return)

    return jsonify({
        "client":      data,
        "last_return": last_return,
        "dependents":  deps,
        "habit_profile": habit_profile,
    })


@app.get("/api/search")
@login_required
def api_search():
    q    = request.args.get("q", "").strip()
    year = int(request.args.get("year", date.today().year))
    if not q:
        return jsonify([])
    results = query_returns({"year": year, "q": q})
    return jsonify([
        {
            "id":         r["id"],
            "log_number": r["log_number"],
            "name":       (f"XXXXX #{r['id']}" if privacy_mode_enabled() else r["name_full"]),
            "status":     r["client_status"],
            "badge":      r["badge_class"],
            "tax_year":   r["tax_year"],
        }
        for r in results[:12]
    ])


@app.post("/api/privacy-mode")
@login_required
def api_privacy_mode():
    data = request.get_json(force=True) if request.data else {}
    enabled = data.get("enabled")
    session["privacy_mode"] = bool(enabled)
    return jsonify({"success": True, "privacy_mode": bool(session.get("privacy_mode"))})


@app.post("/api/return/<int:return_id>/status")
@login_required
def api_status(return_id: int):
    data       = request.get_json(force=True)
    new_status = (data.get("status") or "").upper().strip()
    if new_status not in STATUS_FLOW:
        return jsonify({"error": "Invalid status"}), 400

    conn = get_connection()
    row  = conn.execute("SELECT * FROM returns WHERE id=?", (return_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    old_status  = row["client_status"]
    timestamp   = now()
    today_iso   = date.today().isoformat()
    date_field  = STATUS_DATE_STAMP.get(new_status)

    if date_field and not row[date_field]:
        conn.execute(
            f"UPDATE returns SET client_status=?, {date_field}=?, updated_at=? WHERE id=?",
            (new_status, today_iso, timestamp, return_id),
        )
    else:
        conn.execute(
            "UPDATE returns SET client_status=?, updated_at=? WHERE id=?",
            (new_status, timestamp, return_id),
        )

    # Deduplicated status event
    exists = conn.execute(
        "SELECT id FROM status_events WHERE return_id=? AND event_type='STATUS_CHANGED' AND event_timestamp=?",
        (return_id, timestamp),
    ).fetchone()
    if not exists:
        conn.execute(
            """
            INSERT INTO status_events
              (return_id, event_type, old_status, new_status, event_timestamp, source_file, note)
            VALUES (?, 'STATUS_CHANGED', ?, ?, ?, 'APP', 'Updated via app')
            """,
            (return_id, old_status, new_status, timestamp),
        )

    conn.commit()
    conn.close()
    return jsonify({
        "success":       True,
        "client_status": new_status,
        "badge_class":   STATUS_BADGE.get(new_status, "bg-slate-100 text-slate-500 border-slate-200"),
    })


@app.post("/api/return/<int:return_id>/field")
@login_required
def api_field(return_id: int):
    data  = request.get_json(force=True)
    field = (data.get("field") or "").strip()
    value = data.get("value")

    conn = get_connection()
    try:
        if field in RETURN_EDITABLE:
            conn.execute(
                f"UPDATE returns SET {field}=?, updated_at=? WHERE id=?",
                (value, now(), return_id),
            )
            # Auto-advance to LOG OUT when a completion date is recorded.
            # logout_date = physically logged out; ack_date = IRS accepted.
            # Either one means the engagement is closed.
            auto_logout = (
                (field == "ack_date"    and value) or
                (field == "logout_date" and value)
            )
            if auto_logout:
                cur = conn.execute(
                    "SELECT client_status FROM returns WHERE id=?", (return_id,)
                ).fetchone()
                if cur and cur["client_status"] != "LOG OUT":
                    old_status = cur["client_status"]
                    ts = now()
                    note = "Auto-advanced: ack date set" if field == "ack_date" else "Auto-advanced: logout date set"
                    conn.execute(
                        "UPDATE returns SET client_status='LOG OUT', updated_at=? WHERE id=?",
                        (ts, return_id),
                    )
                    conn.execute(
                        """INSERT INTO status_events
                           (return_id, event_type, old_status, new_status, event_timestamp, source_file, note)
                           VALUES (?, 'STATUS_CHANGED', ?, 'LOG OUT', ?, 'APP', ?)""",
                        (return_id, old_status, ts, note),
                    )
        elif field in CLIENT_EDITABLE:
            cid = conn.execute("SELECT client_id FROM returns WHERE id=?", (return_id,)).fetchone()
            if cid:
                conn.execute(
                    f"UPDATE clients SET {field}=?, updated_at=? WHERE id=?",
                    (value, now(), cid["client_id"]),
                )
        elif field in PAYMENT_EDITABLE:
            prow = conn.execute(
                "SELECT id FROM payments WHERE return_id=?", (return_id,)
            ).fetchone()
            if prow:
                conn.execute(
                    f"UPDATE payments SET {field}=? WHERE return_id=?", (value, return_id)
                )
            else:
                conn.execute(
                    f"INSERT INTO payments (return_id, {field}) VALUES (?,?)", (return_id, value)
                )
        else:
            return jsonify({"error": "Field not editable"}), 400

        conn.commit()
        return jsonify({"success": True, "field": field, "value": value})
    finally:
        conn.close()


@app.post("/api/return/<int:return_id>/note")
@login_required
def api_note(return_id: int):
    data = request.get_json(force=True)
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "Empty note"}), 400

    conn = get_connection()
    dup  = conn.execute(
        "SELECT id FROM notes WHERE return_id=? AND lower(note_text)=lower(?)",
        (return_id, text),
    ).fetchone()
    if dup:
        conn.close()
        return jsonify({"error": "Duplicate note"}), 409

    ts = now()
    conn.execute(
        "INSERT INTO notes (return_id, note_text, source, created_at) VALUES (?,?,'APP',?)",
        (return_id, text, ts),
    )
    conn.commit()
    conn.close()
    visible_text = _mask_value(text) if privacy_mode_enabled() else text
    return jsonify({"success": True, "text": visible_text, "created_at": ts})


# ── Duplicate client detection & merge ───────────────────────────────────────

# For joint clients, "CARLOS G & MARIA" → compare only the primary ("CARLOS G").
def _first_primary_for_compare(first_name: str) -> str:
    s = (first_name or "").upper().strip()
    if " & " in s:
        s = s.split(" & ")[0].strip()
    return s


def _name_tokens(s: str) -> list[str]:
    return _first_primary_for_compare(s).split()


# Extra tokens 1–2 letters (O, A) or Jr/Sr/II, etc. — treat as middle / suffix, same person
_MIDDLE_LIKE: frozenset = frozenset(
    {
        "JR", "SR", "II", "III", "IV", "V",
    }
)


def _token_is_initial_or_suffix(tok: str) -> bool:
    t = tok.strip(".,'").upper()
    if t in _MIDDLE_LIKE:
        return True
    if not t or not t.isalpha():
        return False
    if len(t) == 1:
        return True
    if len(t) == 2 and t.isupper():
        return True
    return False


def _first_names_likely_same_middles(a_first: str, b_first: str) -> bool:
    """
    Same person when the only first-name difference is missing vs middle initial
    (e.g. BRYAN vs BRYAN O, JOSE vs JOSE A, CARLOS vs CARLOS G for one-char 'G').

    If the shorter token list is a prefix of the longer, and every extra token is
    1–2 letter initial/suffix, treat as a duplicate.
    """
    ta, tb = _name_tokens(a_first), _name_tokens(b_first)
    if not ta or not tb:
        return False
    if ta == tb:
        return True
    if len(ta) > len(tb):
        ta, tb = tb, ta
    # ta is shorter
    if len(ta) > len(tb) or not tb:
        return False
    if ta != tb[: len(ta)]:
        return False
    rest = tb[len(ta) :]
    return all(_token_is_initial_or_suffix(x) for x in rest)


def _first_names_likely_duplicate(fa: str, fb: str) -> bool:
    fa_st = (fa or "").upper().strip()
    fb_st = (fb or "").upper().strip()
    if not fa_st or not fb_st:
        return False
    # String containment / one side extends the other (incl. joint "X" in "X & Y")
    if (
        fa_st.startswith(fb_st) or fb_st.startswith(fa_st) or
        fa_st in fb_st or fb_st in fa_st
    ):
        return True
    if _first_names_likely_same_middles(fa, fb):
        return True
    return False


def _find_duplicate_pairs() -> list[dict]:
    """
    Find likely duplicate client records with the same last_name and either:
    - overlapping / contained first_name strings, or
    - first names that differ only by middle initials / extra 1–2 char tokens
      (BRYAN vs BRYAN O) using the primary name before " & " for joint filers.
    """
    conn = get_connection()
    clients = conn.execute(
        """
        SELECT c.id, c.last_name, c.first_name, c.display_name,
               COUNT(r.id)                         AS return_count,
               MAX(r.log_number)                   AS best_log,
               GROUP_CONCAT(r.id)                  AS return_ids,
               GROUP_CONCAT(COALESCE(r.log_number,''))  AS log_numbers,
               GROUP_CONCAT(r.tax_year)            AS tax_years,
               GROUP_CONCAT(r.client_status)       AS statuses
        FROM clients c
        LEFT JOIN returns r ON r.client_id = c.id
        GROUP BY c.id
        ORDER BY c.last_name, c.first_name
        """
    ).fetchall()
    conn.close()

    # Group by last_name
    by_last: dict[str, list] = {}
    for row in clients:
        key = (row["last_name"] or "").upper().strip()
        by_last.setdefault(key, []).append(dict(row))

    pairs = []
    seen = set()
    for last, group in by_last.items():
        if len(group) < 2:
            continue
        for i, a in enumerate(group):
            for b in group[i + 1:]:
                if not (a.get("first_name") or "").strip() or not (b.get("first_name") or "").strip():
                    continue
                if not _first_names_likely_duplicate(
                    a["first_name"] or "", b["first_name"] or ""
                ):
                    continue
                key = tuple(sorted([a["id"], b["id"]]))
                if key in seen:
                    continue
                seen.add(key)
                # Prefer keeping the one with a log number / more returns
                a_score = (1 if a["best_log"] else 0) + (a["return_count"] or 0)
                b_score = (1 if b["best_log"] else 0) + (b["return_count"] or 0)
                keep, discard = (a, b) if a_score >= b_score else (b, a)
                pairs.append({
                    "keep":    keep,
                    "discard": discard,
                })

    return sorted(pairs, key=lambda p: (p["keep"]["last_name"] or ""))


def _merge_pair_key(ka: int, kb: int) -> str:
    return f"{min(ka, kb)}-{max(ka, kb)}"


def _merge_pairs_for_session() -> list[dict]:
    skipped = set(session.get("merge_skipped", []))
    out: list[dict] = []
    for p in _find_duplicate_pairs():
        k, d = int(p["keep"]["id"]), int(p["discard"]["id"])
        if _merge_pair_key(k, d) in skipped:
            continue
        out.append(p)
    return out


@app.get("/merge-clients")
@login_required
def merge_clients_page():
    pairs = _merge_pairs_for_session()
    ctx = base_ctx(date.today().year)
    ctx.update({"active_page": "merge", "pairs": pairs})
    return render_template("merge_clients.html", **ctx)


@app.post("/api/merge-clients")
@login_required
def api_merge_clients():
    """
    Merge 'discard' client into 'keep' client.
    Moves all returns (and review_queue refs) from discard → keep, then deletes discard.
    """
    data       = request.get_json(force=True)
    keep_id    = int(data.get("keep_id", 0))
    discard_id = int(data.get("discard_id", 0))
    if not keep_id or not discard_id or keep_id == discard_id:
        return jsonify({"error": "Invalid IDs"}), 400

    conn = get_connection()
    try:
        keep    = conn.execute("SELECT * FROM clients WHERE id=?", (keep_id,)).fetchone()
        discard = conn.execute("SELECT * FROM clients WHERE id=?", (discard_id,)).fetchone()
        if not keep or not discard:
            return jsonify({"error": "Client not found"}), 404

        ts = now()
        merge_client_into(conn, keep_id, discard_id, ts)
        conn.commit()

        keep_name = (keep["display_name"] or
                     f"{keep['last_name']}, {keep['first_name']}".strip(", "))
        return jsonify({"success": True, "kept": keep_name})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.post("/api/merge-clients/bulk")
@login_required
def api_merge_clients_bulk():
    """
    Auto-merge all duplicate pairs (same as Merge Dupes list, respecting skipped pairs in session).
    `dry_run: true` returns a count and sample; false runs the merge in separate transactions.
    """
    data    = request.get_json(silent=True) or {}
    dry_run = bool(data.get("dry_run"))
    limit   = int(data.get("limit", 2000))
    if limit < 1 or limit > 5000:
        limit = 2000

    pairs = _merge_pairs_for_session()[:limit]
    if dry_run:
        return jsonify({
            "success":  True,
            "dry_run":  True,
            "count":    len(pairs),
            "previews": [
                {
                    "keep_id":    int(p["keep"]["id"]),
                    "discard_id": int(p["discard"]["id"]),
                    "name": (p["keep"]["last_name"] or "")
                    + ", " + (p["keep"].get("first_name") or ""),
                }
                for p in pairs[:50]
            ],
        })

    merged = 0
    errors: list[dict] = []
    for p in pairs:
        k = int(p["keep"]["id"])
        d = int(p["discard"]["id"])
        conn = get_connection()
        try:
            merge_client_into(conn, k, d, now())
            conn.commit()
            merged += 1
        except Exception as e:
            conn.rollback()
            errors.append({
                "keep_id":    k, "discard_id": d, "error": str(e),
            })
        finally:
            conn.close()

    return jsonify({"success": True, "merged": merged, "errors": errors})


@app.post("/api/merge-clients/skip")
@login_required
def api_merge_skip():
    """Mark a pair as 'not duplicates' by storing a skip record (simple session list)."""
    data = request.get_json(force=True)
    skipped = session.get("merge_skipped", [])
    pair_key = f"{min(data['keep_id'], data['discard_id'])}-{max(data['keep_id'], data['discard_id'])}"
    if pair_key not in skipped:
        skipped.append(pair_key)
    session["merge_skipped"] = skipped
    return jsonify({"success": True})


# ── Startup ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    conn = get_connection()
    init_db(conn)
    conn.close()
    print("TaxOps running at http://localhost:5000")
    app.run(debug=True, host="0.0.0.0", port=5000)
